#!/usr/bin/env python3
"""Compact structural QA for local spreadsheet deliverables."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from zipfile import BadZipFile, ZipFile

import openpyxl
import pandas as pd


ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}


def inspect_workbook(file_path: Path) -> tuple[dict[str, object], list[str]]:
    issues: list[str] = []
    try:
        with ZipFile(file_path) as archive:
            corrupt_member = archive.testzip()
            if corrupt_member:
                issues.append(f"corrupt archive member: {corrupt_member}")
    except BadZipFile:
        return {"kind": "workbook", "sheets": []}, ["not a valid OOXML workbook"]

    keep_vba = file_path.suffix.lower() == ".xlsm"
    workbook = openpyxl.load_workbook(file_path, data_only=False, read_only=False, keep_vba=keep_vba)
    sheets: list[dict[str, object]] = []
    for sheet in workbook.worksheets:
        formula_count = 0
        error_cells: list[str] = []
        nonempty_count = 0
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                nonempty_count += 1
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                if value in ERROR_VALUES:
                    error_cells.append(cell.coordinate)
        if nonempty_count == 0:
            issues.append(f"empty sheet: {sheet.title}")
        if error_cells:
            issues.append(f"formula error literals in {sheet.title}: {', '.join(error_cells[:20])}")
        sheets.append({
            "name": sheet.title,
            "rows": sheet.max_row,
            "columns": sheet.max_column,
            "nonemptyCells": nonempty_count,
            "formulas": formula_count,
            "mergedRanges": len(sheet.merged_cells.ranges),
            "hidden": sheet.sheet_state != "visible",
        })
    workbook.close()
    return {"kind": "workbook", "sheets": sheets}, issues


def inspect_delimited(file_path: Path) -> tuple[dict[str, object], list[str]]:
    separator = "\t" if file_path.suffix.lower() == ".tsv" else ","
    frame = pd.read_csv(file_path, sep=separator)
    issues = ["table has no rows"] if frame.empty else []
    return {
        "kind": "table",
        "rows": len(frame),
        "columns": len(frame.columns),
        "columnNames": [str(column) for column in frame.columns],
        "nullCells": int(frame.isna().sum().sum()),
        "duplicateRows": int(frame.duplicated().sum()),
    }, issues


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("file", type=Path)
    args = parser.parse_args()
    file_path = args.file.resolve()
    if not file_path.is_file():
        parser.error(f"file not found: {file_path}")
    suffix = file_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        summary, issues = inspect_workbook(file_path)
    elif suffix in {".csv", ".tsv"}:
        summary, issues = inspect_delimited(file_path)
    else:
        parser.error(f"unsupported spreadsheet type: {suffix}")
    result = {"file": file_path.name, **summary, "issues": issues, "ok": not issues}
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if not issues else 1


if __name__ == "__main__":
    raise SystemExit(main())
